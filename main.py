import pandas as pd
import openpyxl as pyxl
import re
import sys


def update_excel(file_path):
    """
    输出文件
    :param path: 文件输出路径
    :return:
    """
    # 获取时间分类字典
    time_dict = create_time_dict(file_path)
    # 打开excel
    print("\n打开文件中...")
    wb = pyxl.load_workbook(filename=file_path)
    # 打开sheet
    ws = wb['自动分类']
    # 目标数据范围
    col_range = "A1:A" + str(ws.max_row)
    items = ws[col_range]
    for item in items:
        # 实现进度条
        i = round((item[0].row/ws.max_row)*100)
        print("\r", end="")
        print("完成进度: {}%: ".format(i), "▋" * (i // 2), end="")
        sys.stdout.flush()

        # 获取分类信息
        if item[0].value :
            my_classify = find_class(item[0].value,time_dict=time_dict)
            if my_classify:
                # print(my_classify)
                ws["B" + str(item[0].row)] = my_classify['大类']
                ws["C" + str(item[0].row)] = my_classify['小类']
                ws["D" + str(item[0].row)] = my_classify['明细类']
    print("\n保存中...")
    wb.save(filename=file_path)



def create_time_dict(file_path):
    """
    创建时间分类字典
    :param file_path:
    :return:
    """
    dict_df = pd.read_excel(file_path, sheet_name='分类字典')
    my_dict = []
    # 制作一个二维矩阵来作为字典方便检索
    my_dict.append(list(dict_df['大类']))
    my_dict.append(list(dict_df['小类']))
    my_dict.append(list(dict_df['明细类']))
    my_dict.append(list(dict_df['规则']))

    # 清洗空值
    for j in range(len(my_dict[3])):
        if not isinstance(my_dict[3][j], str):
            my_dict[3][j] = None

    for j in range(len(my_dict[2])):
        if not isinstance(my_dict[2][j], str):
            my_dict[2][j] = None

    for j in range(len(my_dict[1])):
        if not isinstance(my_dict[1][j], str):
            my_dict[1][j] = None

    for j in range(len(my_dict[0])):
        if not isinstance(my_dict[0][j], str):
            my_dict[0][j] = None
    return my_dict


def find_class(record,time_dict):
    """
    检索时间分类
    :param records:
    :return:
    """
    find_result = {}
    # 获取时间分类字典
    for i in range(len(time_dict[2])):
        regular = time_dict[3][i]
        # 先使用正则匹配
        if regular:
            # print(time_dict[3][i])
            pattern = re.compile(time_dict[3][i])
            try:
                re_result = pattern.findall(record)
            except Exception as e:
                print(e)
            if re_result:
                find_result['data'] = record
                find_result['大类'] = time_dict[0][i]
                find_result['小类'] = time_dict[1][i]
                find_result['明细类'] = time_dict[2][i]
        # 再使用明细类覆盖
        if record == time_dict[2][i]:
            find_result['data'] = record
            find_result['大类'] = time_dict[0][i]
            find_result['小类'] = time_dict[1][i]
            find_result['明细类'] = time_dict[2][i]
    return find_result


if __name__ == '__main__':
    # file_path = r"/Users/bancine/OneDrive/Documents/时间统计法原表邦胜.xlsx"
    file_path = sys.argv[1]
    print("执行开始".center(50 // 2, "-"))
    update_excel(file_path)
    print("\n")
    print("执行结束".center(50 // 2, "-"))


